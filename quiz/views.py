from django.db.models import Count, Q
import openpyxl.cell
from .models import PerfilVisitante, VisitantePerguntaResposta
from perfil_visitante.models import PerfilVisitante
from django.contrib import messages

from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import Pergunta, Resposta
from .forms import PerguntaForm, RespostaForm
import calendar
import datetime
from django.db.models import Count
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from io import BytesIO




def criar_pergunta(request):
    if request.method == 'POST':
        form = PerguntaForm(request.POST)
        if form.is_valid():
            pergunta = form.save()      
            messages.success(request, "Pergunta criada com sucesso!")  
            return redirect('criar_resposta', pergunta_id=pergunta.id)
    else:
        form = PerguntaForm() 
    perguntas = Pergunta.objects.all()   
    return render(request, 'criar_pergunta.html', {'form': form, 'peguntas': perguntas})


def editar_pergunta(request, pergunta_id):
    pergunta = get_object_or_404(Pergunta, id=pergunta_id)    
    if request.method == 'POST':
        form = PerguntaForm(request.POST, instance=pergunta)
        if form.is_valid():
            form.save()
            messages.success(request,  'Pergunta atualizada com sucesso!')
            return redirect('criar_pergunta')
    else:
        form = PerguntaForm(instance=pergunta)
    
    perguntas = Pergunta.objects.all()   
    return render(request, 'criar_pergunta.html', {'form': form, 'peguntas': perguntas})


def deletar_pergunta(request, pergunta_id):
    pergunta = get_object_or_404(Pergunta, id=pergunta_id)
    pergunta.delete()
    messages.success(request, 'Pergunta excluída com sucesso')
    return redirect('criar_pergunta')


# RESPOSTAS

from django.urls import reverse_lazy
def criar_resposta(request, pergunta_id):
    if request.method == 'POST':
        form = RespostaForm(request.POST)
        if form.is_valid():            
            form.instance.pergunta = Pergunta.objects.get(id = pergunta_id)
            form.save()     
            messages.success(request, 'Resposta criada com sucesso')       
            return redirect(reverse_lazy('criar_resposta', kwargs={'pergunta_id': pergunta_id}))
   
    else:
        form = RespostaForm()   
    pergunta_p = Pergunta.objects.get(id = pergunta_id)
    resposta_r = Resposta.objects.filter(pergunta = pergunta_id) 
    return render(request, 'criar_resposta.html', {'form': form, 'pergunta': pergunta_p, 'resposta':resposta_r})

def editar_resposta(request, resposta_id):
    resposta = get_object_or_404(Resposta, id=resposta_id)
    
    if request.method == 'POST':
        form = RespostaForm(request.POST, instance=resposta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Resposta atualizada com sucesso!!')
            return redirect('listar_respostas')
    else:
        form = RespostaForm(instance=resposta)
    
    return render(request, 'editar_resposta.html', {'form': form, 'resposta': resposta})


def deletar_resposta(request, resposta_id):
    # Obtém a resposta a ser deletada
    resposta = get_object_or_404(Resposta, id=resposta_id)

    # Deleta a resposta
    resposta.delete()

    # Obtém o ID da pergunta para redirecionamento
    pergunta_id = resposta.pergunta.id

    # Retorna para a página de criação de resposta
    messages.success(request, 'Resposta excluída com sucesso')
    return redirect('criar_resposta', pergunta_id=pergunta_id)


def relatorio_completo(request):
    if request.method == 'GET':
        mes_str = request.GET.get('mes')
        try:
            mes = int(mes_str) if mes_str and mes_str.isdigit() else None
        except ValueError:
            mes = None
        
        if mes:
            resposta_visitante = VisitantePerguntaResposta.objects.filter(criado_em__month=mes)
            perfil_visitante = PerfilVisitante.objects.filter(criado_em__month=mes)
            mes_nome = calendar.month_name[mes]
        else:
            resposta_visitante = VisitantePerguntaResposta.objects.all()
            perfil_visitante = PerfilVisitante.objects.all()
            mes_nome = "Todos os meses"

        # Quantidade de visitantes
        total_visitantes = perfil_visitante.count()

        # Quantidade de respostas corretas e incorretas
        total_acertos = resposta_visitante.filter(resposta__correta=True).count()
        total_erros = resposta_visitante.filter(resposta__correta=False).count()

        # Respostas mais acertadas
        respostas_acertadas = resposta_visitante.values('resposta__texto_resposta').annotate(total=Count('id')).order_by('-total')

        # Cidades que mais acertaram
        cidades_acertos = perfil_visitante.values('municipio_escola').annotate(total_acertos=Count(
            'Visitante_related', filter=Q(Visitante_related__resposta__correta=True))).order_by('-total_acertos')

        # Notas mais dadas
        notas_mais_dadas = perfil_visitante.values('nota_visita').annotate(total=Count('id')).order_by('-total')

        # Idades mais visitadas
        idades_mais_visitadas = perfil_visitante.values('idade').annotate(total=Count('id')).order_by('-total')

        # Quantidade de estudantes por turma
        quant_estudante_turma = perfil_visitante.values('turma').annotate(total=Count('id')).order_by('turma')

        # Mapear os valores da turma para seus nomes completos
        turma_mapping = dict(PerfilVisitante.TURMA_CHOICES)
        for turma in quant_estudante_turma:
            turma['turma'] = turma_mapping.get(turma['turma'], turma['turma'])

        context = {
            'total_visitantes': total_visitantes,
            'total_acertos': total_acertos,
            'total_erros': total_erros,
            'respostas_acertadas': respostas_acertadas,
            'cidades_acertos': cidades_acertos,
            'notas_mais_dadas': notas_mais_dadas,
            'idades_mais_visitadas': idades_mais_visitadas,
            'quant_estudante_turma': quant_estudante_turma,
            'mes_nome': mes_nome,
            'mes': mes
        }

        if request.GET.get('export') == 'excel':
            messages.success(request, 'Relatório exportado com sucesso')
            response = exportar_para_excel(context)
            return response
        else:
            return render(request, 'relatorio_completo.html', context)


def exportar_para_excel(context):
    # Criação de uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Completo"

    # Estilos
    title_font = Font(size=14, bold=True)    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    data_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    alignment = Alignment(horizontal="center", vertical="center")

    
     # Configuração do papel
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # Tamanho do papel A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # Orientação retrato

    # Configuração das margens
    ws.page_margins = PageMargins(
        left=0.75,
        right=1,
        top=1.0,
        bottom=1.0,
        header=0.3,
        footer=0.3
    )
    
    # Função para aplicar estilos a uma linha
    def style_row(row):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    # Função para adicionar e formatar cabeçalhos e dados
    def add_section(title, rows):
        headers = [title, 'Valor']  # Cabeçalhos para seções de dados
        ws.append(headers)
        style_row(ws[ws.max_row])
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
        
        for row in rows:
            ws.append(row)
            style_row(ws[ws.max_row])
        
        ws.append([])  # Linha em branco após os dados

    # Adicionando Total Visitantes, Total Acertos, Total Erros
    
    meses = {
    "January": 'Janeiro',
    "February": 'Fevereiro',
    "March": 'Março',
    "April": 'Abril',
    "May": 'Maio',
    "June": 'Junho',
    "July": 'Julho',
    "August": 'Agosto',
    "September": 'Setembro',
    "October": 'Outubro',
    "November": 'Novembro',
    "December": 'Dezembro'
    }
    mes_nome = context['mes_nome']
    mes = meses.get(mes_nome, "(Soma de todos os meses)")

    
    #Adicionando título geral da planilha
    titulo = f"Resumo de Todas as Visitas de Alunos:AA {mes}"
    ws.merge_cells('A1:B1') # Mesclar as celulas da primeira linha
    ws['A1'] = titulo
    ws['A1'].font = title_font
    ws['A1'].alignment = alignment

    # Liha em branco após o titulo
    ws.append([])
    
    #Adicionando título geral da planilha
    titulo = f"Resumo de Todas as Visitas de Alunos: {mes}"
    ws.merge_cells('A1:B1') # Mesclar as celulas da primeira linha
    ws['A1'] = titulo
    ws['A1'].font = title_font
    ws['A1'].alignment = alignment

    # Liha em branco após o titulo
    ws.append([])


    add_section('Resumo Geral', [
        ['Periodo do ano', mes],
        ['Total Visitantes', context['total_visitantes']],
        ['Total Acertos', context['total_acertos']],
        ['Total Erros', context['total_erros']]
    ])

     

    # Adicionando Respostas Acertadas
    add_section('Respostas Acertadas', [
        [resposta['resposta__texto_resposta'], resposta['total']]
        for resposta in context['respostas_acertadas']
    ])

    # Adicionando Cidades Acertos
    add_section('Cidades Acertos', [
        [cidade['municipio_escola'], cidade['total_acertos']]
        for cidade in context['cidades_acertos']
    ])

    # Adicionando Notas Mais Dadas
    add_section('Notas Mais Dadas', [
        [nota['nota_visita'], nota['total']]
        for nota in context['notas_mais_dadas']
    ])

    # Adicionando Idades Mais Visitadas
    add_section('Idades Mais Visitadas', [
        [idade['idade'], idade['total']]
        for idade in context['idades_mais_visitadas']
    ])

    # Adicionando Quantidade de Estudantes por Turma
    add_section('Quantidade de Estudantes por Turma', [
        [turma['turma'], turma['total']]
        for turma in context['quant_estudante_turma']
    ])

    # Definindo a largura da segunda coluna 
    ws.column_dimensions['B'].width =30

    # Ajustando largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter if isinstance(col[0], openpyxl.cell.cell.Cell) else None
        if column: # Apenas ajustar se a coluna não for celula mesclada
            for cell in col:
                try:
                    if isinstance(cell, openpyxl.cell.cell.Cell) and  len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    ws.column_dimensions['A'].width =50

    # Ajustar o alinhamento da coluna A para esquerda
    for cell in ws['A']:
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Adicionando uma linha com texto informativo ao final
    ws.append([])
    informativo_text = "Este relatório é gerado automaticamente e contém informações detalhadas sobre as visitas de alunos durante o período mencionado"
    ws.append([informativo_text])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
    ws[ws.max_row][0].alignment = Alignment(
        horizontal="center",
        vertical="center", 
        wrap_text=True,
        )
    
    # Area de impressão
    #ws.print_area = 'A1:B20'

    # Criando o buffer de memória e salvando a planilha nele
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Criando a resposta HTTP
    response = HttpResponse(content=buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_completo.xlsx"'
    return response
